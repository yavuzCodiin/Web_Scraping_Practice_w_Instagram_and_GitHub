from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import time
import openpyxl
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import NoSuchElementException
from bs4 import BeautifulSoup

browser = webdriver.Chrome()
actions = ActionChains(browser)

username_text = "Username"
password_text = "Password"

#Initialize Excel Workbook
wb = openpyxl.Workbook()
ws = wb.active

# Column Headers
ws.append(['Followers', 'Following', 'Non-Followers'])

browser.get("https://www.instagram.com/")

time.sleep(2)

username = browser.find_element(By.NAME, "username")
password = browser.find_element(By.NAME, "password")

username.send_keys(username_text)
time.sleep(5)
password.send_keys(password_text)
time.sleep(3)
actions.send_keys(Keys.RETURN)
actions.perform()
time.sleep(10)

save_info = browser.find_element(By.TAG_NAME, "button")
save_info.click()

time.sleep(20)

# Using XPath to find the button by its text
turn_on_button = browser.find_element(By.XPATH, "//button[text()='Not Now']")
turn_on_button.click()

time.sleep(5)

browser.get("https://www.instagram.com/{}/followers/".format(username_text))
time.sleep(15)

#The following variable is for scrolling down 
jscommand = """
followers = document.querySelector("._aano");
followers.scrollTo(0, followers.scrollHeight);
var lenOfPage=followers.scrollHeight;
return lenOfPage;
"""

lenOfPage = browser.execute_script(jscommand)

match=False
while(match==False):
    lastCount = lenOfPage
    time.sleep(1)
    lenOfPage = browser.execute_script(jscommand)
    if lastCount == lenOfPage:
        match=True
time.sleep(5)

followersList = []

html_content = browser.page_source
soup = BeautifulSoup(html_content, 'html.parser')

follower_elements = soup.find_all('span', class_='_ap3a _aaco _aacw _aacx _aad7 _aade')

for follower in follower_elements:
    followersList.append(follower.text)
print("Followers: {}".format(followersList))

browser.get("https://www.instagram.com/{}/following/".format(username_text))
time.sleep(15)
lenOfPage = browser.execute_script(jscommand)
match=False
while(match==False):
    lastCount = lenOfPage
    time.sleep(1)
    lenOfPage = browser.execute_script(jscommand)
    if lastCount == lenOfPage:
        match=True
time.sleep(5)

html_content = browser.page_source
soup = BeautifulSoup(html_content, 'html.parser')

following_elements = soup.find_all('span', class_='_ap3a _aaco _aacw _aacx _aad7 _aade')

for following in following_elements:
    followingList.append(following.text)
print("Following: {}".format(followingList))

follows = set(followersList)
following = set(followingList)
not_follower = following.difference(follows)

# Writing to Excel

# Write Followers
row = 1  # Start from the second row (first row is for headers)
for follower in followersList:
    ws.cell(row=row, column=1, value=follower)
    row += 1

# Write Following
row = 1  # Reset row for following
for follow in followingList:
    ws.cell(row=row, column=2, value=follow)
    row += 1

# Write Non-Followers
row = 1  # Reset row for non-followers
for non_follow in not_follower:
    ws.cell(row=row, column=3, value=non_follow)
    row += 1

wb.save('non_follower.xlsx')     
          
browser.close()





























